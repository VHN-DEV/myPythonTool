#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tool: Xuất báo cáo commits GitLab ra Excel

Mục đích: Export commits với thông tin chi tiết, phân loại tăng ca
Lý do: Tracking công việc, báo cáo tăng ca, quản lý dự án
"""

import os
import sys
import re
import subprocess
from pathlib import Path
from datetime import datetime, time
from collections import defaultdict

# Thêm thư mục cha vào sys.path để import utils
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from utils import (
    print_header, get_user_input, confirm_action,
    log_info, log_error, normalize_path, install_library
)
from utils.colors import Colors

# Kiểm tra thư viện
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.borders import Border, Side
except ImportError:
    install_library(
        package_name="openpyxl",
        install_command="pip install openpyxl",
        library_display_name="openpyxl"
    )
    sys.exit(1)


def cut_after_pipe(s):
    """Cắt chuỗi lấy phần trước dấu '|'"""
    return s.split('|', 1)[0].strip()


def extract_op_id(message):
    """Tìm OP id trong message dạng {OP#1234}"""
    m = re.search(r"\{OP#(\d+)\}", message)
    return m.group(1) if m else None


def get_branch_for_hash(commit_hash):
    """Lấy branch chứa commit"""
    try:
        result = subprocess.run(
            ["git", "branch", "--contains", commit_hash, "--format=%(refname:short)"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, 
            universal_newlines=True, check=True, encoding='utf-8'
        )
        branches = [line.strip() for line in result.stdout.splitlines() if line.strip()]
        return ", ".join(branches)
    except Exception:
        return ""


def get_child_commits_by_author(parent, merge_hash, author):
    """Lấy danh sách commit con trong merge commit"""
    try:
        result = subprocess.run(
            ["git", "log", f"{parent}..{merge_hash}", "--pretty=format:%h|%an", f"--author={author}"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            universal_newlines=True, check=True, encoding='utf-8'
        )
        lines = result.stdout.strip().splitlines()
        hashes = [line.split("|")[0] for line in lines if line.strip()]
        return hashes
    except Exception:
        return []


def get_git_commits(since, until, author=None, branch=None, keyword=None):
    """Lấy danh sách commits từ Git"""
    git_cmd = ["git", "log"]
    
    if branch:
        git_cmd.append(branch)
    else:
        git_cmd.append("--all")
    
    git_cmd += [
        f"--since={since}",
        f"--until={until}",
        '--pretty=format:%ad\x1f%H\x1f%an\x1f%D\x1f%s\x1f%P',
        "--date=format:%Y-%m-%d %H:%M:%S"
    ]
    
    if author:
        git_cmd.append(f"--author={author}")
    
    try:
        result = subprocess.run(
            git_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            universal_newlines=True, encoding='utf-8'
        )
        
        if result.returncode != 0:
            return []
        
        entries = []
        for line in result.stdout.strip().splitlines():
            parts = line.strip().split("\x1f")
            if len(parts) == 6:
                date, hash_, author_name, ref, msg, parents = parts
                entries.append({
                    "date": date.strip(),
                    "hash": hash_.strip(),
                    "author": author_name.strip(),
                    "ref": ref.strip(),
                    "msg": msg.strip(),
                    "parents": parents.strip()
                })
        
        # Lọc theo keyword nếu có
        if keyword:
            entries = [
                e for e in entries
                if any(kw.lower() in e["msg"].lower() for kw in keyword)
            ]
        
        return entries
    except Exception as e:
        log_error(f"Lỗi khi lấy commits: {e}")
        return []


def process_commits(entries, start_work=time(8, 0), end_work=time(17, 30), include_working_hours=True):
    """Xử lý và phân loại commits"""
    data = defaultdict(lambda: defaultdict(list))
    count_stat = defaultdict(lambda: {"overtime": 0, "working": 0})
    
    for entry in entries:
        try:
            dt = datetime.strptime(entry["date"], "%Y-%m-%d %H:%M:%S")
            is_overtime = not (start_work <= dt.time() <= end_work)
            
            if not include_working_hours and not is_overtime:
                continue
            
            author = cut_after_pipe(entry["author"])
            commit_hash = entry["hash"].strip()
            ref_info = entry["ref"].strip()
            parents = entry.get("parents", "").strip().split()
            is_merge = len(parents) > 1
            
            # Lấy commit con của merge
            child_commits = []
            if is_merge and len(parents) > 1:
                child_commits = get_child_commits_by_author(parents[1], commit_hash, author)
            
            # Xác định type
            commit_type = (
                "merge-with-contribution" if is_merge and child_commits else
                "merge-only" if is_merge else
                "commit"
            )
            
            # Lấy branch
            branch = ""
            m = re.search(r'HEAD -> ([^,\s]+)', ref_info)
            if m:
                branch = m.group(1)
            elif "origin/" in ref_info:
                m = re.search(r'origin/([^,\s]+)', ref_info)
                if m:
                    branch = m.group(1)
            elif ref_info:
                branch = ref_info.split(",")[0].strip()
            
            if not branch:
                branch = get_branch_for_hash(commit_hash)
            
            branch = re.sub(r"^origin/", "", branch)
            
            # Lưu dữ liệu
            data[author][dt.date()].append({
                "time": dt.strftime("%H:%M"),
                "hash": commit_hash,
                "message": entry["msg"].strip(),
                "branch": branch,
                "event": "merge" if is_merge else "commit",
                "type": commit_type,
                "child_commits": child_commits,
                "is_overtime": is_overtime
            })
            
            # Cập nhật thống kê
            if is_overtime:
                count_stat[author]["overtime"] += 1
            else:
                count_stat[author]["working"] += 1
                
        except Exception as e:
            log_error(f"Lỗi xử lý commit {entry.get('hash', 'unknown')}: {e}")
            continue
    
    return data, count_stat


def export_to_excel(data, count_stat, output_path, include_working_hours=True):
    """Export dữ liệu ra Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tăng ca"
    
    headers = ["Tác giả", "Ngày", "Thời gian", "Mã commit", "Event", "Type", "Branch", "Nội dung", "OP Link", "Child Commits"]
    ws.append(headers)
    
    # Style
    header_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    summary_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFC07F", end_color="FFC07F", fill_type="solid")
    alt_row_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Header style
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.freeze_panes = "A2"
    row = 2
    first_author = True
    
    for author, dates in data.items():
        if not first_author:
            ws.append([""] * len(headers))
            row += 1
        else:
            first_author = False
        
        author_start_row = row
        
        for date, commits in sorted(dates.items()):
            for c in commits:
                op_id = extract_op_id(c["message"])
                row_data = [
                    author,
                    date.strftime("%d/%m/%Y"),
                    c["time"],
                    c["hash"],
                    c["event"],
                    c["type"],
                    c["branch"],
                    c["message"],
                    op_id if op_id else "",
                    ", ".join(c["child_commits"]) if c["child_commits"] else ""
                ]
                ws.append(row_data)
                
                # Style rows
                if row % 2 == 0:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col_idx).fill = alt_row_fill
                
                # Hyperlink OP
                if op_id:
                    ws.cell(row=row, column=9).hyperlink = f"https://work.fsofts.com/work_packages/{op_id}"
                    ws.cell(row=row, column=9).style = "Hyperlink"
                
                # Alignment
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).alignment = Alignment(vertical="top", horizontal="left")
                
                # Bold merge
                if c["event"] == "merge":
                    for col_idx in (5, 6):
                        ws.cell(row=row, column=col_idx).font = Font(bold=True)
                
                # Highlight overtime
                if c["is_overtime"]:
                    ws.cell(row=row, column=3).fill = highlight_fill
                
                row += 1
        
        # Author name ở dòng đầu
        for r in range(author_start_row, row):
            cell = ws.cell(r, 1)
            if r == author_start_row:
                cell.value = author
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(r, col_idx).border = Border(top=Side(style='medium'))
            else:
                cell.value = ""
        
        # Summary row
        o = count_stat[author]["overtime"]
        w = count_stat[author]["working"]
        summary = f"Tăng ca: {o} commits"
        if include_working_hours:
            summary += f" | Trong giờ: {w} commits"
        
        ws.append([f"Tổng cộng: {summary}"])
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=10)
        merged_cell = ws.cell(row=row, column=1)
        merged_cell.alignment = Alignment(horizontal="left", vertical="top")
        merged_cell.fill = summary_fill
        merged_cell.font = Font(bold=True)
        row += 1
    
    # Auto width
    MAX_WIDTH = 30
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[1:]:
            try:
                cell_value = str(cell.value or "")
            except:
                cell_value = ""
            if len(cell_value) > max_len:
                max_len = len(cell_value)
        max_len = min(max_len, MAX_WIDTH)
        ws.column_dimensions[col_letter].width = max(max_len + 1, 8)
    
    wb.save(output_path)
    log_info(f"✅ Đã xuất Excel: {output_path}")


def print_summary(data, count_stat, include_working_hours):
    """In tóm tắt ra console"""
    for author, dates in data.items():
        print(f"\n{Colors.primary('='*60)}")
        print(f"{Colors.bold(f'👤 Tác giả: {author}')}")
        print(Colors.primary('='*60))
        
        for date, commits in sorted(dates.items()):
            print(f"📅 Ngày: {date.strftime('%d/%m/%Y')}")
            for c in commits:
                print(f"  🕒 {c['time']} - 🔗 {c['hash'][:8]} | {c['type']} | [{c['branch']}] | {c['message'][:50]}...")
                if c['child_commits']:
                    print(f"     ↳ Gồm commit con: {', '.join(c['child_commits'])}")
        
        o = count_stat[author]["overtime"]
        w = count_stat[author]["working"]
        summary = f"🔥 Tăng ca: {o} commit(s)"
        if include_working_hours:
            summary += f" | ⏰ Trong giờ: {w} commit(s)"
        print(f"\n📊 {summary}")


def main():
    """Hàm main"""
    print_header()
    print(Colors.primary("  📊 TOOL XUẤT BÁO CÁO COMMITS GITLAB"))
    print("=" * 70)
    print()
    
    # Kiểm tra Git repo
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--git-dir"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE
        )
        if result.returncode != 0:
            print(Colors.error("❌ Không phải Git repository!"))
            print("Hãy chạy tool trong thư mục Git repository.")
            return 1
    except Exception:
        print(Colors.error("❌ Git không được cài đặt hoặc không tìm thấy!"))
        return 1
    
    # Nhập thông tin
    since = get_user_input("Từ ngày (YYYY-MM-DD hoặc YYYY-MM-DDTHH:MM:SS): ", default="2025-01-01T00:00:00")
    until = get_user_input("Đến ngày (YYYY-MM-DD hoặc YYYY-MM-DDTHH:MM:SS): ", default="2025-01-31T23:59:59")
    author = get_user_input("Tác giả (Enter để bỏ qua): ", default="")
    branch = get_user_input("Branch (Enter để tất cả): ", default="")
    keyword_input = get_user_input("Từ khóa trong message (cách nhau bằng dấu phẩy, Enter để bỏ qua): ", default="")
    keyword = [k.strip() for k in keyword_input.split(',')] if keyword_input else None
    
    include_working = confirm_action("Bao gồm commit trong giờ làm việc?", default=False)
    
    # Lấy commits
    print(f"\n🔍 Đang lấy commits...")
    entries = get_git_commits(since, until, author if author else None, 
                              branch if branch else None, keyword)
    
    if not entries:
        print(Colors.warning("⚠️  Không tìm thấy commit nào!"))
        return 0
    
    print(Colors.success(f"✅ Tìm thấy {len(entries)} commit(s)"))
    
    # Xử lý
    print(f"\n📊 Đang xử lý và phân loại...")
    data, count_stat = process_commits(entries, include_working_hours=include_working)
    
    if not data:
        print(Colors.warning("⚠️  Không có dữ liệu để xuất!"))
        return 0
    
    # In summary
    print_summary(data, count_stat, include_working)
    
    # Export Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(f"commits_{timestamp}.xlsx")
    
    if confirm_action(f"\n💾 Xuất ra Excel ({output_path})?", default=True):
        export_to_excel(data, count_stat, output_path, include_working)
        print(Colors.success(f"📄 File Excel: {output_path}"))
    
    print()
    print(Colors.success("✅ Hoàn tất!"))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print(Colors.warning("\n⚠️  Đã hủy bởi người dùng!"))
        sys.exit(130)
    except Exception as e:
        log_error(f"❌ Lỗi không mong muốn: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

